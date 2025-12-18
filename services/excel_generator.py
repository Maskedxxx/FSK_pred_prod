"""Сервис генерации Excel отчёта по дефектам.

Принимает результат дедупликации и создаёт форматированный Excel файл
с колонками для всех полей дефектов и пометками дубликатов.

Публичный API:
    - generate_excel_report() — основная функция генерации
"""

from __future__ import annotations

import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import logger
from services.defect_deduplicator import DeduplicationResult


# =============================================================================
# Настройки стилей Excel
# =============================================================================

# Заголовок таблицы
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Ячейки данных
DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="top")

# Дубликаты (подсветка)
DUPLICATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Границы
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Ширина колонок
COLUMN_WIDTHS = {
    "A": 5,    # №
    "B": 8,    # Страница
    "C": 15,   # Помещение
    "D": 18,   # Локализация
    "E": 30,   # Тип дефекта
    "F": 20,   # Тип работы
    "G": 80,   # Описание
    "H": 12,   # Дубликаты
}

# Заголовки колонок
HEADERS = [
    "№",
    "Стр.",
    "Помещение",
    "Локализация",
    "Тип дефекта",
    "Тип работы",
    "Описание дефекта",
    "Дубликаты",
]


# =============================================================================
# Основная логика
# =============================================================================


def generate_excel_report(
    dedup_result: DeduplicationResult,
    output_path: str | Path | None = None,
    output_dir: str | Path = "artifacts/excel",
) -> str:
    """Генерирует Excel отчёт по дефектам.

    Args:
        dedup_result: Результат дедупликации дефектов
        output_path: Полный путь к выходному файлу (опционально)
        output_dir: Папка для сохранения (используется если output_path не указан)

    Returns:
        Путь к сохранённому Excel файлу
    """
    start_time = time.perf_counter()

    # Определяем путь к файлу
    if output_path:
        excel_path = Path(output_path).expanduser().resolve()
    else:
        out_dir = Path(output_dir).expanduser().resolve()
        out_dir.mkdir(parents=True, exist_ok=True)
        pdf_stem = Path(dedup_result.source_pdf).stem
        excel_path = out_dir / f"defects_{pdf_stem}.xlsx"

    logger.info(
        "Генерация Excel: %d дефектов → %s",
        dedup_result.total_defects,
        excel_path,
    )

    # Создаём workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефекты"

    # Заголовки
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Устанавливаем ширину колонок
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Фиксируем заголовок
    ws.freeze_panes = "A2"

    # Данные
    for row_idx, defect in enumerate(dedup_result.defects, start=2):
        row_data = [
            defect.row_number,
            defect.page_number,
            defect.room,
            defect.location,
            defect.defect,
            defect.work_type,
            defect.source_text,
            defect.duplicates_str,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            # Выравнивание
            if col_idx in (1, 2, 8):  # №, Страница, Дубликаты
                cell.alignment = DATA_ALIGNMENT_CENTER
            else:
                cell.alignment = DATA_ALIGNMENT

            # Подсветка дубликатов
            if defect.has_duplicates:
                cell.fill = DUPLICATE_FILL

    # Высота строк (авто не работает, ставим минимум для wrap_text)
    for row_idx in range(2, len(dedup_result.defects) + 2):
        ws.row_dimensions[row_idx].height = 60

    # Сохраняем
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)

    elapsed = time.perf_counter() - start_time

    logger.info(
        "Excel сохранён: %s (%.2f сек)",
        excel_path,
        elapsed,
    )

    return str(excel_path)
